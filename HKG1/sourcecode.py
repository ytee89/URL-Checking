# -*- coding: utf-8 -*-
"""
Created on Tue Jul  3 11:45:04 2018

@author: zmohamadazri
"""
import os
import openpyxl
import xlrd
import time
from packagetools.urlaccess import URL, deletefile

def checkupdate(url, indicator_name, stpname, save_path, driver, timepoint1, ref):
    u = URL(url)
    
    if url.endswith(".pdf"):
        u.dlfile(save_path)
        if indicator_name == "OFCA":
            file_size = os.path.getsize(save_path)
            last_update = str(file_size) + " byte"
            deletefile(save_path)
        else:
            last_update = u.pdfmoddate(save_path)
            deletefile(save_path)
            
    else:
        if indicator_name == "HKICL":
            savepath = save_path.replace(save_path[save_path.rfind('.'):], '.pdf')
            u.dlfile(savepath)
            last_update = u.pdfmoddate(savepath)
            deletefile(savepath)
            
        elif indicator_name == "CHP":
            u.dlfile(save_path)
           
            while not os.path.exists(save_path):
                time.sleep(1)                 
            if os.path.isfile(save_path):
                wwb = openpyxl.load_workbook(save_path)
                sheete = wwb.active
                rw = 4
                while sheete.cell(row=rw,column=4).value != None:
                    rw += 1
                last_update = str(sheete.cell(row=rw-1,column=4).value)
            else:
                raise ValueError("%s not found!" % save_path)
            
            deletefile(save_path)
                
        elif indicator_name == "GEM - b":
            u.dlfile(save_path)
            
            while not os.path.exists(save_path):
                time.sleep(1)                 
            if os.path.isfile(save_path):
                wwb = xlrd.open_workbook(save_path)
                sheete = wwb.sheet_by_index(0)
                cl = 3
                while sheete.cell(rowx=11,colx=cl).value != "":
                    cl += 1
                last_update = str(sheete.cell(rowx=8,colx=cl-1).value)
            else:
                raise ValueError("%s not found!" % save_path)
                
            deletefile(save_path)
            
        elif indicator_name == "HKMA(MSB) - 03":
            u.dlfile(save_path)
            
            while not os.path.exists(save_path):
                time.sleep(1)                 
            if os.path.isfile(save_path):
                wwb = xlrd.open_workbook(save_path)
                sheete = wwb.sheet_by_index(0)
                rw = 3
                while sheete.cell(rowx=rw,colx=1).value != "":
                    rw += 1
                last_update = str(sheete.cell(rowx=rw-1,colx=0).value)
            else:
                raise ValueError("%s not found!" % save_path)
                
            deletefile(save_path)
                
        elif indicator_name == "HKMA(MSB) - 06":
            u.dlfile(save_path)
            
            while not os.path.exists(save_path):
                time.sleep(1)                 
            if os.path.isfile(save_path):
                wwb = xlrd.open_workbook(save_path)
                sheete = wwb.sheet_by_index(0)
                rw = 3
                while sheete.cell(rowx=rw,colx=2).value != "":
                    rw += 1
                last_update = str(sheete.cell(rowx=rw-1,colx=1).value)
            else:
                raise ValueError("%s not found!" % save_path)
                
            deletefile(save_path)
                
        elif indicator_name == "HKMA":
            last_update1 = u.urlrequests().find(id='lastUpdate').text
            last_update = last_update1[2 + last_update1.find(":") - len(last_update1):]
            
        elif indicator_name == "HKMA1":
            for last_update1 in u.urlrequests().findAll('tr'):
                if stpname in last_update1.text:
                    last_update = last_update1.find('td').text
                    break
        
        elif indicator_name == "GEM - a":
            last_update1 = u.urlrequests().find('span', attrs={'class': 'header'}).text
            last_update = last_update1[2 + last_update1.find("-") - len(last_update1):]
        
        elif indicator_name == "HKMC":
            last_update1 = u.urlrequests().find('span', attrs={'class':'date'}).text
            last_update = last_update1[1 + last_update1.find("(") - len(last_update1):]
        
        elif indicator_name == "TRY":
            last_update = u.urlrequests().find('a', attrs={'class':'text_link'}).text
        
        elif indicator_name == "HKMA(MSB)":
            for last_update1 in u.urlrequests().findAll('tr'):
                if last_update1.text.startswith(stpname):
                    last_update2 = last_update1.findAll('td')[-1].text
                    last_update = last_update2[:6]
                
        elif indicator_name == "THB":
            last_update1 = u.urlrequests().findAll(id='psContent')
            for div in last_update1:
                last_update = div.find('a')['href']
            
        elif indicator_name == "RVD":
            last_update = u.urlrequests().find(id='article').find('h1').text
        
        elif indicator_name == "BD":
            for last_update1 in u.urlrequests().findAll('td'):
                if last_update1.text.startswith(ref):
                    last_update2 = last_update1.find('span', attrs={'class':'updated_on'})
                    last_update = last_update2.text
                
        elif indicator_name == "HA":
            for last_update1 in u.urlrequests().findAll('a'):
                if last_update1.text.startswith('Annual Report'):
                    last_update = last_update1.text
                
        elif indicator_name == "CAD":
            last_update = u.urlrequests().findAll('strong')[-1].text
        
        elif indicator_name == "COL":
            for last_update1 in u.urlrequests().findAll('a'):
                if 'Hong Kong Office Market' in last_update1.text:
                    last_update = last_update1.text[:7]
                    
        elif indicator_name == "COSCO":
            for last_update1 in u.urlrequests().findAll('div', attrs={'class':'text'}):
                if 'Container Throughput' in last_update1.text:
                    last_update = last_update1.text
                
        elif indicator_name == "HKEX - a":
            last_update = u.urlrequests().find('p', attrs={'class':'loadMore__timetag'}).text
        
        elif indicator_name == "HKIA":
            last_update = u.urlrequests().find('td', attrs={'class':'month'}).text
        
        elif indicator_name == "IRD":
            last_update = u.urlrequests().findAll('strong')[0].text
        
        elif indicator_name == "LAD":
            last_update1 = u.urlrequests().find('td', attrs={'class':'review_dt_txt'}).text
            last_update = last_update1[2 + last_update1.find(":") - len(last_update1):]
        
        elif indicator_name =="LR":
            last_update1 = u.urlrequests().find('title').text
            last_update = last_update1[4 + last_update1.find("for") - len(last_update1):]

        elif indicator_name == "MPFA":
            last_update1 = u.urlrequests().find('p', attrs={'class':'reviewdate'}).text
            last_update = last_update1[2 + last_update1.find(":") - len(last_update1):]
        
        elif indicator_name == "MTR - a":
            last_update1 = u.urlrequests().find('div', attrs={'class':'table_title'}).text
            last_update = last_update1[4 + last_update1.find("for") - len(last_update1):]
        
        elif indicator_name == "OCI":
            last_update = u.urlrequests().findAll('li')[0].text
            
        elif indicator_name == "ID - a":
            for last_update1 in u.urlrequests().findAll('div', attrs={'class':'showHide'}):
                if last_update1.text.startswith(stpname):
                    last_update = last_update1.findAll('th')[-1].text
                    
        elif indicator_name == "ID - b":
            for last_update1 in u.urlrequests().findAll('div', attrs={'class':'showHide'}):
                if last_update1.text.startswith(stpname):
                    last_update2 = last_update1.findAll('p')[1].text
                    last_update = last_update2[last_update2.rfind("(") - len(last_update2):]
                    
        elif indicator_name == "SFC":
            last_update1 = u.urlrequests().find('p', attrs={'class': 'lastmodified'})
            last_update = last_update1.text[2 + last_update1.text.find(":") - len(last_update1.text):]
            
        elif indicator_name == "CENTALINE1":
            last_update1 = u.urlrequests().find(id='_ctl2_myDataGrid')
            for last_update2 in last_update1.findAll('tr'):
                if stpname in last_update2.text:
                    last_update = last_update2.text[-11:]
                    
        elif indicator_name == "HKE":
            for last_update1 in u.urlrequests().findAll('p'):
                if last_update1.text.startswith("Economic Situation and GDP and Price Forecasts"):
                    last_update = last_update1.text[last_update1.text.find("(") - len(last_update1.text):]
                    
            try:
                last_update
            except NameError:
                last_update = str(timepoint1)
                    
        elif indicator_name == "CENTALINE2":
            for last_update1 in u.urlrequests().findAll('div', attrs={'class':'articleMoreItem'}):
                if last_update1.text.startswith("\n" + ref):
                    last_update = last_update1.text[2 + last_update1.text.rfind(":") - len(last_update1.text):]
                    break
            try:
                last_update
            except NameError:
                last_update = str(timepoint1)
                
        elif indicator_name == "CENTALINE3":
            for last_update1 in u.urlrequests().findAll('div', attrs={'class':'articleMoreItem'}):
                if 'CCL' in last_update1.text:
                    last_update2 = last_update1.findAll('span')[1].text
                    last_update = last_update2[2 + last_update2.rfind(":") - len(last_update2):]
                    break
                
        elif indicator_name == "HKMAD":
            last_update = u.urlrequests().find('td', attrs={'class':'left date'}).text
            
        elif indicator_name == "HKAB":
            for last_update1 in u.urlrequests().findAll('td'):
                if last_update1.text.startswith('Last updated:'):
                    last_update2 = last_update1.text[2 + last_update1.text.find(":") - len(last_update1.text):]
                    last_update = last_update2.strip()
                    
        elif indicator_name == "CPAL1":
            for last_update1 in u.urlrequests().findAll('b'):
                if 'Cathay Pacific' and 'Traffic Figures' in last_update1.text:
                    last_update = last_update1.text
                                  
        elif indicator_name == "HKEX1":
            last_update = u.urlrequests().find(id='select-target-firstdropdownlist').text
            
        elif indicator_name == "CPAL2":
            for last_update1 in u.urlrequests().findAll('tbody'):
                if 'Annual Reports' in last_update1.text:
                    last_update = last_update1.findAll('tr')[1].text
                    
    
        elif indicator_name == "CUHK":
            last_update1 = u.urlrequests().find(id='t5')
            last_update2 = last_update1.findAll('tr')[0]
            last_update = last_update2.findAll('td')[-1].text
            
        elif indicator_name == "GEMD":
            last_update1 = u.urlrequests().find(id='ctl00_ContentPlaceHolder1_hrefDailyQuotation')['href']
            last_update = last_update1[3 + last_update1.rfind("/") - len(last_update1):]
            
        elif indicator_name == "ISD":
            for last_update1 in u.urlrequests().findAll('a'):
                if last_update1.text.startswith('Effective Exchange Rate Index'):
                    last_update2 = last_update1['href']
                    last_update = last_update2[1 + last_update2.rfind("/") - len(last_update2):]
            try:
                last_update
            except NameError:
                last_update = str(timepoint1)
                
        elif indicator_name == "SCMP":
            last_update1 = u.urlrequests().find('div',attrs={'class':'epaper-main-wrapper'})
            last_update = last_update1.find('div',attrs={'class':'date'}).text
            
        elif indicator_name == "HKEXD":
            try:
                last_update1 = u.urlrequests().find('div',attrs={'class':'static-calendar'})
                last_update = last_update1.findAll('a', href=True)[-1].text
            
            except IndexError as iderror:
                last_update1 = u.urlrequests().findAll('div',attrs={'class':'static-calendar'})[-1]
                last_update = last_update1.findAll('a', href=True)[-1].text
            
        elif indicator_name == "HKAB2":
            last_update = u.urlrequests().find('b').text
            
        elif indicator_name == "HKEXD1":
            last_update1 = u.urlrequests().find('font').text
            last_update = last_update1[:126][115:]
            
        elif indicator_name == "HKEXD2":
            last_update1 = u.urlrequests().find('font').text
            last_update = last_update1[:119][108:]
            
        elif indicator_name == "IDDS":
            for last_update1 in u.urlrequests().findAll('li'):
                if last_update1.text.startswith('Statistics updated to'):
                    last_update = last_update1.text[3 + last_update1.text.rfind("to") - len(last_update1.text):]
                    
        elif indicator_name == "LD":
            last_update1 = u.urlrequests().find('table')
            last_update = last_update1.findAll('a')[-1].text.strip()
            
        elif indicator_name == "HSI":
            last_update1 = u.urlgetjson()
            lalala = []

            for lala in last_update1['indexSeriesList'][0]['reportList'][2]['reportDate'][:]:
                la = lala['date'][:10].strip()
                lalala.append(la)
            
            last_update = ','.join(lalala)
                                
        elif indicator_name == "HKEX2":
            last_update1 = u.urlgetjson()['tables'][0]['body']
            last_update2 = last_update1[-7]['text']
            
            if '(' in last_update2:
                last_update = str(last_update1[-14]['text'])
            else:
                last_update = str(last_update2)
                
        elif indicator_name == "ID2":
            for last_update1 in u.urlrequests().findAll('a',href=True):
                if last_update1.text.startswith\
                ('Special Meetings of the Finance Committee to Examine the Draft Estimates of Expenditure'):
                    last_update = last_update1.text[-7:]
                    
        elif indicator_name == "UGC":
            last_update = u.urlrequests().findAll('text')[-1].text
            
        elif indicator_name == "HKEXD3":
            last_update1 = u.urlgetjson()
            last_update = str(last_update1['tables'][0]['body'][-7]['text'])
            
        elif indicator_name == "CR":
            last_update1 = u.urlrequests().find(id='main').find('table')
            for last_update2 in last_update1.findAll('tr')[3:-1]:
                last_update3 = last_update2.findAll('td')[-1]
                if not last_update3.text.isspace():
                    last_update = last_update2.find('td').text
                    
        elif indicator_name == "ORO":
            last_update = u.urlrequests().find('select', attrs={'name':'end_year'})['onchange'][-9:-2].strip()
            
        elif indicator_name == "TD":
            last_update = u.urlrequests().find(id='tagHead').find('meta',attrs={'name':'date'})['content']
            
        elif indicator_name == "LR2":
            last_update = u.urlrequests().find('meta',attrs={'name':'date'})['content']
            
        elif indicator_name == "LD2":
            last_update1 = u.urlrequests().find('head').findAll('script')[-1].text
            last_update2 = last_update1[last_update1.find('revisionDate'):]
            last_update = last_update2[:last_update2.find(';')].strip()
            
        elif indicator_name == "RVD2":
            last_update = u.urlrequests().find('head').findAll('script')[-1].text.strip()
            
        elif indicator_name == "TD":
            last_update = u.urlrequests().find('meta', attrs={'name':'date'})['content']
            
        elif indicator_name == "HKTB1":
            driver.set_page_load_timeout(30)
            driver.get(url)
            driver.find_element_by_xpath('//*[@id="mainForm:subFormLogin:txtLoginName"]').send_keys('hkceic39')
            driver.find_element_by_xpath('//*[@id="mainForm:subFormLogin:txtPassword"]').send_keys('isiceic123')
            driver.find_element_by_xpath('//*[@id="mainForm:subFormLogin:btnLogin"]').click()
            
            la = u.getdriver(driver).find('ul' , attrs={'class':'accordion'}).find('ul', attrs={'class':'multi_col col_4 clearfix'})
            lala = la.find('li').text
            last_update = lala[:lala.find('\n')].strip()
                    
        elif indicator_name == "HKTB2":
            la = u.getdriver(driver).find('ul' , attrs={'class':'accordion'}).find('ul', attrs={'class':'multi_col col_4 clearfix'})
            lala = la.find('li').text
            last_update = lala[:lala.find('\n')].strip()
                    
        elif indicator_name == "HKTB3":
            la = u.getdriver(driver).find('ul' , attrs={'class':'accordion'})
            last_update = la.find('div', attrs={'class':'list_header'}).text.strip()
            
        elif indicator_name == "HKTB4":
            last_update = u.getdriver(driver).find(id='gvHelper:gvHelper_dataTabl:tbody_element').find('td').text
            
        elif indicator_name == "HKPC":
            la = u.urlrequests().find('div', attrs={'itemprop':'articleBody'}).find('table').find('a').text
            last_update = la[2+la.find('-'):la.find('(')].strip()
            
    return last_update