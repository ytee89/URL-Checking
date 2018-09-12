# -*- coding: utf-8 -*-
"""
Created on Tue Jul 10 12:38:29 2018

@author: zmohamadazri
"""

import win32com.client as win32
import pandas as pd
import openpyxl
             
def dftomasterfile(excelfile, df):
    #write dataframe to excel masterfile
    wb = openpyxl.load_workbook(excelfile)
    for row in wb['Sheet1']:
        for cell in row:
            cell.value = None
    writer = pd.ExcelWriter(excelfile,engine='openpyxl')
    writer.book = wb
    writer.sheets = dict((ws.title,ws) for ws in wb.worksheets)
    df.to_excel(writer,sheet_name='Sheet1',index=False)
    writer.save()
    
def dftoreport(excelfile, df):
    #write dataframe to excel report attachment
    wb = openpyxl.Workbook(excelfile)
    wb.save(excelfile)
    wb1 = openpyxl.load_workbook(excelfile)
    
    writer = pd.ExcelWriter(excelfile,engine='openpyxl')
    writer.book = wb1
    writer.sheets = dict((ws.title,ws) for ws in wb1.worksheets)
    df.to_excel(writer,sheet_name='Sheet',index=False)
    writer.save()
    
def exceltohtml(excelfile, htmlfile):
    #convert excel to html
    excel = win32.DispatchEx('Excel.Application')
    wb = excel.Workbooks.Open(excelfile)
#    wb.Visible = False
    ob = wb.PublishObjects.Add(1, htmlfile, 'Sheet')
    ob.Publish(True)
    wb.Save()
    wb.Close()

