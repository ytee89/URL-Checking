# -*- coding: utf-8 -*-
"""
Created on Tue Jul 10 11:48:49 2018

@author: zmohamadazri
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

class CheckingResult(object):
    def __init__(self, i, cursor, df2, systemid):
        self.i = i#variable for the mdb iteration
        self.cursor = cursor#mdb connection from the masterfile
        self.df2 = df2#dataframe from the report attachment
        self.systemid = systemid
        
    def failed(self, result):
        #Change status to failed in in the mdb
        self.cursor.execute("UPDATE URLChecking SET Remark = '" + result + "' WHERE SystemID = '" + self.systemid + "';")
        
        self.df2.loc[self.i, 'Source'] = self.cursor.execute("SELECT Source from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'STP'] = self.cursor.execute("SELECT STPName from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'New Timepoint'] = self.cursor.execute("SELECT TimePointSource from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'Previous Timepoint'] = self.cursor.execute("SELECT LastTimepoint from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'Remark'] = self.cursor.execute("SELECT Remark from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'Key'] = self.cursor.execute("SELECT KeySeries from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'Frequency'] = self.cursor.execute("SELECT Frequency from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'Level'] = self.cursor.execute("SELECT PublicationLevel from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'System ID'] = self.cursor.execute("SELECT SystemID from URLChecking").fetchall()[self.i][0]
        
    def updatedetected(self):
        #Change status to new detected in the mdb
        self.cursor.execute("UPDATE URLChecking SET LastTimepoint = CurrentTimePoint  WHERE SystemID = '" + self.systemid + "';")
        self.cursor.execute("UPDATE URLChecking SET CurrentTimePoint = TimePointSource  WHERE SystemID = '" + self.systemid + "';")
        self.cursor.execute("UPDATE URLChecking SET Remark = 'New Detected'  WHERE SystemID = '" + self.systemid + "';")
        self.cursor.execute("UPDATE URLChecking SET Status = 'Done'  WHERE SystemID = '" + self.systemid + "';")
        
        self.df2.loc[self.i, 'Source'] = self.cursor.execute("SELECT Source from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'STP'] = self.cursor.execute("SELECT STPName from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'New Timepoint'] = self.cursor.execute("SELECT TimePointSource from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'Previous Timepoint'] = self.cursor.execute("SELECT LastTimepoint from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'Remark'] = self.cursor.execute("SELECT Remark from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'Key'] = self.cursor.execute("SELECT KeySeries from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'Frequency'] = self.cursor.execute("SELECT Frequency from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'Level'] = self.cursor.execute("SELECT PublicationLevel from URLChecking").fetchall()[self.i][0]
        self.df2.loc[self.i, 'System ID'] = self.cursor.execute("SELECT SystemID from URLChecking").fetchall()[self.i][0]
        
    def uptodate(self):  
        #Change status to up to date in the mdb
        self.cursor.execute("UPDATE URLChecking SET Remark = 'Up to date'  WHERE SystemID = '" + self.systemid + "';")
        self.cursor.execute("UPDATE URLChecking SET Status = 'Done'  WHERE SystemID = '" + self.systemid + "';")

class ExcelChanges(object):
    def __init__(self, excelfile2, excelfile3, cursor):
        wb = openpyxl.Workbook(excelfile3)
        wb.save(excelfile3)
        
        self.excelfile3 = excelfile3#excel file for email body
        self.excelfile2 = excelfile2#report attachment excel file
        self.cursor = cursor#mdb connection from the masterfile
        
    def reporttoemail(self):
        #write email body to the excel file before html conversion
        wb2 = openpyxl.load_workbook(self.excelfile2)
        sheet2 = wb2['Sheet']
        wb3 = openpyxl.load_workbook(self.excelfile3)
        sheet3 = wb3['Sheet']
        
        sheet2.column_dimensions['A'].width = 13
        sheet2.column_dimensions['B'].width = 36
        sheet2.column_dimensions['C'].width = 25
        sheet2.column_dimensions['D'].width = 25
        sheet2.column_dimensions['E'].width = 23
        sheet2.column_dimensions['F'].width = 4
        sheet2.column_dimensions['G'].width = 11
        sheet2.column_dimensions['H'].width = 12
        sheet2.column_dimensions['I'].width = 12
        
        sheet3.cell(row=1, column=1).value = 'Source'
        sheet3.cell(row=1, column=2).value = 'STP'
        sheet3.cell(row=1, column=3).value = 'Remark'
        sheet3.cell(row=1, column=4).value = 'Key'
        sheet3.cell(row=1, column=5).value = 'Frequency'
        sheet3.cell(row=1, column=6).value = 'Level'
        sheet3.cell(row=1, column=7).value = 'System ID'
        
        sheet3.column_dimensions['A'].width = 15
        sheet3.column_dimensions['B'].width = 38
        sheet3.column_dimensions['C'].width = 23
        sheet3.column_dimensions['D'].width = 5
        sheet3.column_dimensions['E'].width = 14
        sheet3.column_dimensions['F'].width = 12
        sheet3.column_dimensions['G'].width = 10
        
        for colnum in range(1,8):
            sheet3.cell(row=1,column=colnum).font = Font(bold=True)
            sheet3.cell(row=1,column=colnum).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                                              right = Side(border_style='thin', color='FF000000'), 
                                                              bottom = Side(border_style='thin', color='FF000000'),
                                                              left = Side(border_style='thin', color='FF000000'))
        
        i = 2
        
        while sheet2.cell(row=i,column=2).value != None:
            for la in range(len(self.cursor.execute("SELECT * from URLChecking").fetchall())):
                if sheet2.cell(row=i, column=9).value == self.cursor.execute("SELECT SystemID from URLChecking").fetchall()[la][0]:
                    sheet2.cell(row=i, column=2).hyperlink = self.cursor.execute("SELECT RealURL from URLChecking").fetchall()[la][0]
                    
            sheet3.cell(row=i,column=1).value = sheet2.cell(row=i,column=1).value
            sheet3.cell(row=i,column=2).value = sheet2.cell(row=i,column=2).value
            sheet3.cell(row=i,column=2).font = Font(color = "0000FF")
            sheet2.cell(row=i,column=2).font = Font(color = "0000FF")
            sheet3.cell(row=i,column=2).hyperlink = sheet2.cell(row=i,column=2).hyperlink
            sheet3.cell(row=i,column=3).value = sheet2.cell(row=i,column=5).value
            sheet3.cell(row=i,column=4).value = sheet2.cell(row=i,column=6).value
            sheet3.cell(row=i,column=5).value = sheet2.cell(row=i,column=7).value
            sheet3.cell(row=i,column=6).value = sheet2.cell(row=i,column=8).value
            sheet3.cell(row=i,column=7).value = sheet2.cell(row=i,column=9).value
            
            if sheet2.cell(row=i,column=5).value == "New Detected":
                sheet2.cell(row=i,column=3).font = Font(color="FFFF0000")
                for colNum in range(1,10):
                    sheet2.cell(row=i,column=colNum).fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
                for colNum in range(1,8):
                    sheet3.cell(row=i,column=colNum).fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
            
            for colnum in range(1,8):
                sheet3.cell(row=i,column=colnum).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                                                  right = Side(border_style='thin', color='FF000000'), 
                                                                  bottom = Side(border_style='thin', color='FF000000'),
                                                                  left = Side(border_style='thin', color='FF000000'))
                
            for colnum in range(1,10):
                sheet2.cell(row=i,column=colnum).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                                                  right = Side(border_style='thin', color='FF000000'), 
                                                                  bottom = Side(border_style='thin', color='FF000000'),
                                                                  left = Side(border_style='thin', color='FF000000'))
                
                
            i += 1
            
        wb2.save(self.excelfile2)
        wb3.save(self.excelfile3)