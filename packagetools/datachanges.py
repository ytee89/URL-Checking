# -*- coding: utf-8 -*-
"""
Created on Tue Jul 10 11:48:49 2018

@author: zmohamadazri
"""
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import pyodbc
import datetime

def consolidate(df2, consofile):
    if not os.path.exists(consofile):
        wb = openpyxl.Workbook(consofile)
        wb.save(consofile)
        df3 = pd.DataFrame(columns=['Source','STP Name', 'New Timepoint','Previous Timepoint', 'Changes Type', 
                                    'Key', 'Frequency', 'Level', 'System ID', 'Method', 'Remark', 'Requested Time'])
    else:
        df3 = pd.read_excel(consofile)
        
    for i in df2.index:
        if df2.loc[i, 'Changes Type'] == 'New Detected':
            j = len(df3)
            df3.loc[j] = df2.loc[i]
    
    wb1 = openpyxl.load_workbook(consofile)
    writer = pd.ExcelWriter(consofile,engine='openpyxl')
    writer.book = wb1
    writer.sheets = dict((ws.title,ws) for ws in wb1.worksheets)
    df3.to_excel(writer,sheet_name='Sheet',index=False)
    writer.save()
    
class CheckingResult(object):
    def __init__(self, i, df1, df2):
        self.i = i#variable for the dataframe iteration
        self.df1 = df1#dataframe from the masterfile
        self.df2 = df2#dataframe from the report attachment
        
    def copyvalues(self):
        self.df2.loc[self.i, 'Source'] = self.df1.loc[self.i, 'Source']
        self.df2.loc[self.i, 'STP Name'] = self.df1.loc[self.i, 'STP Name']
        self.df2.loc[self.i, 'New Timepoint'] = self.df1.loc[self.i, 'TimePoint Source']
        self.df2.loc[self.i, 'Previous Timepoint'] = self.df1.loc[self.i, 'Last Timepoint']
        self.df2.loc[self.i, 'Changes Type'] = self.df1.loc[self.i, 'Changes Type']
        self.df2.loc[self.i, 'Key'] = self.df1.loc[self.i, 'Key Series']
        self.df2.loc[self.i, 'Frequency'] = self.df1.loc[self.i, 'Frequency']
        self.df2.loc[self.i, 'Level'] = self.df1.loc[self.i, 'Level']
        self.df2.loc[self.i, 'System ID'] = self.df1.loc[self.i, 'System ID']
        self.df2.loc[self.i, 'Method'] = self.df1.loc[self.i, 'Update Method']
        self.df2.loc[self.i, 'Remark'] = self.df1.loc[self.i, 'Remark']
        self.df2.loc[self.i, 'Requested Time'] = ((datetime.datetime.now()+datetime.timedelta(hours=8)).strftime('%d-%m-%Y %I:%M:%S %p'))
        
    def failed(self, result):
        #Change status to failed in in the dataframe
        self.df1.loc[self.i, 'Changes Type'] = result
        self.copyvalues()
        
    def updatedetected(self):
        #Change status to new detected in the dataframe
        self.df1.loc[self.i, 'Last Timepoint'] = self.df1.loc[self.i, 'Current TimePoint']
        self.df1.loc[self.i, 'Current TimePoint'] = self.df1.loc[self.i, 'TimePoint Source']
        self.df1.loc[self.i, 'Changes Type'] = "New Detected"
        self.df1.loc[self.i, 'Status'] = "Done"
        self.copyvalues()
        
    def uptodate(self):  
        #Change status to up to date in the dataframe
        self.df1.loc[self.i, 'Changes Type'] = "Up to date"
        self.df1.loc[self.i, 'Status'] = "Done"
        
    def updatemdb(self, mdbfile, countrycode):
        mdbpath = (
        r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
        r"DBQ="+mdbfile+";"
        )
        
        cnxn = pyodbc.connect(mdbpath, autocommit=True)
        cursor = cnxn.cursor()
        
        if self.df1.loc[self.i, 'Key Series'] == 'Y':
            keyseries = 'TRUE'
        else:
            keyseries = 'FALSE'
            
        cursor.execute("INSERT INTO automation (country, publication, source_id, key_dataset, triggered_by, requested_time) VALUES ('"+
                       countrycode+"', '"+self.df1.loc[self.i, 'STP Name']+"', '"+str(self.df1.loc[self.i, 'System ID'])+"', "+
                       keyseries +", 'RC', #"+datetime.datetime.now().strftime("%Y-%b-%d  %H:%M:%S")+"#);")
        
        cursor.close()
        del cursor
        cnxn.close()
        
class ExcelChanges(object):
    def __init__(self, excelfile2, excelfile3, df1):
        wb = openpyxl.Workbook(excelfile3)
        wb.save(excelfile3)
        
        self.excelfile3 = excelfile3#excel file for email body
        self.excelfile2 = excelfile2#report attachment excel file
        self.df1 = df1#dataframe from the masterfile
        
    def reporttoemail(self):
        #write email body to the excel file before html conversion
        wb2 = openpyxl.load_workbook(self.excelfile2)
        sheet2 = wb2['Sheet']
        wb3 = openpyxl.load_workbook(self.excelfile3)
        sheet3 = wb3['Sheet']
        
        for dim1 in [['A',8],['B',36],['C',25],['D',25],['E',23],['F',8],['G',11],['H',12],['I',12],['J',8],['K',25]]:
            sheet2.column_dimensions[dim1[0]].width = dim1[1]
        
        for colname in [[1,'Source'],[2,'STP Name'],[3,'Changes Type'],[4,'Key'],[5,'Frequency'],
                        [6,'Level'],[7,'System ID'],[8,'Method'],[9,'Remark']]:
            sheet3.cell(row=1, column=colname[0]).value = colname[1]
        
        for dim2 in [['A',8],['B',35],['C',21],['D',6],['E',10],['F',11],['G',10],['H',8],['I',22]]:
            sheet3.column_dimensions[dim2[0]].width = dim2[1]
        
        for colnum in range(1,10):
            sheet3.cell(row=1,column=colnum).font = Font(bold=True)
            sheet3.cell(row=1,column=colnum).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                                              right = Side(border_style='thin', color='FF000000'), 
                                                              bottom = Side(border_style='thin', color='FF000000'),
                                                              left = Side(border_style='thin', color='FF000000'))
        
        i = 2
        
        while sheet2.cell(row=i,column=2).value != None:
            for la in self.df1.index:
                if sheet2.cell(row=i, column=9).value == self.df1.loc[la, 'System ID']:
                    sheet2.cell(row=i, column=2).hyperlink = self.df1.loc[la, 'Real URL']
                    
            sheet3.cell(row=i,column=1).value = sheet2.cell(row=i,column=1).value
            sheet3.cell(row=i,column=2).value = sheet2.cell(row=i,column=2).value
            sheet3.cell(row=i,column=2).font = Font(color = "0000FF")
            sheet2.cell(row=i,column=2).font = Font(color = "0000FF")
            sheet3.cell(row=i,column=2).hyperlink = sheet2.cell(row=i,column=2).hyperlink
            sheet3.cell(row=i,column=3).value = sheet2.cell(row=i,column=5).value
            sheet3.cell(row=i,column=4).value = sheet2.cell(row=i,column=6).value
            sheet3.cell(row=i,column=5).value = sheet2.cell(row=i,column=7).value
            sheet3.cell(row=i,column=6).value = sheet2.cell(row=i,column=8).value
            sheet3.cell(row=i,column=7).value = sheet2.cell(row=i,column=9).value
            sheet3.cell(row=i,column=8).value = sheet2.cell(row=i,column=10).value
            sheet3.cell(row=i,column=9).value = sheet2.cell(row=i,column=11).value
            
            if sheet2.cell(row=i,column=5).value == "New Detected":
                sheet2.cell(row=i,column=3).font = Font(color="FFFF0000")
                for colNum in range(1,12):
                    sheet2.cell(row=i,column=colNum).fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
                for colNum in range(1,10):
                    sheet3.cell(row=i,column=colNum).fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
            
            for colnum in range(1,10):
                sheet3.cell(row=i,column=colnum).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                                                  right = Side(border_style='thin', color='FF000000'), 
                                                                  bottom = Side(border_style='thin', color='FF000000'),
                                                                  left = Side(border_style='thin', color='FF000000'))
                
            for colnum in range(1,12):
                sheet2.cell(row=i,column=colnum).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                                                  right = Side(border_style='thin', color='FF000000'), 
                                                                  bottom = Side(border_style='thin', color='FF000000'),
                                                                  left = Side(border_style='thin', color='FF000000'))
                
            i += 1
            
        wb2.save(self.excelfile2)
        wb3.save(self.excelfile3)